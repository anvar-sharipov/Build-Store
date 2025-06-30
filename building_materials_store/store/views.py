from django.shortcuts import render
import time
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from icecream import ic


from rest_framework import viewsets, status, filters
from .models import *
from .serializers import *

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, BasePermission, IsAuthenticated, SAFE_METHODS
from rest_framework.decorators import api_view, action

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import render, get_object_or_404
from rest_framework.generics import CreateAPIView
from django_filters.rest_framework import DjangoFilterBackend
from .filters import ProductFilter
from django.views.decorators.http import require_GET
from django.http import JsonResponse
from django.contrib.postgres.search import TrigramSimilarity
from django.db.models import Q
from django.utils.dateparse import parse_datetime, parse_date
from django.db.models import Sum, F, Count
from openpyxl.styles import Font

from rest_framework.pagination import PageNumberPagination
# swoy pagination 
class CustomPageNumberPagination(PageNumberPagination):
    page_size = 2              # Кол-во элементов на странице по умолчанию
    page_size_query_param = 'page_size'  # Позволяет клиенту указать page_size в запросе (?page_size=50)
    max_page_size = 100          # Максимальное кол-во элементов на странице, чтобы не перегружать сервер



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_user(request):
    user = request.user
    return Response({
        "username": user.username,
        "photo": request.build_absolute_uri(user.photo.url) if user.photo else None
    })


# swoyo razgranichenie dostupa
class IsInAdminOrWarehouseGroup(BasePermission):
    """
    Чтение — всем авторизованным.
    Изменение — только группам 'admin' и 'warehouse_manager'.
    """
    def has_permission(self, request, view):
        user = request.user
        # Сначала проверяем – аутентификация выполнена?
        if not user or not user.is_authenticated:
            return False
        # Разрешаем чтение всем аутентифицированным
        if request.method in SAFE_METHODS:
            return True # GET, HEAD, OPTIONS
        # Для мутации — проверяем группы пользователя
        allowed = user.groups.filter(name__in=['admin', 'warehouse_manager']).exists()
        return allowed





class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer









# стал полнофункциональным ModelViewSet (CRUD),
# работал с одной точки входа (/api/products/),
# использовал групповое разграничение доступа:
# GET, HEAD, OPTIONS → всем авторизованным,
# POST, PUT, PATCH, DELETE → только тем, кто в группе admin или warehouse_manager.

class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    pagination_class = CustomPageNumberPagination

    filter_backends = [DjangoFilterBackend]
    filterset_class = ProductFilter

    def get_permissions(self):
        return [IsInAdminOrWarehouseGroup()]

    def get_queryset(self):
        qs = Product.objects.all()
        qs = qs.select_related('category', 'base_unit', 'brand', 'model')
        qs = qs.prefetch_related('units__unit', 'images', 'batches')
        return qs.distinct()

    def list(self, request, *args, **kwargs):
        # time.sleep(1)  # для теста задержка
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        # time.sleep(1)  # задержка для теста
        return super().retrieve(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        time.sleep(2)
        return super().update(request, *args, **kwargs)








class PartnerViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Partner.objects.all().order_by('-pk')
    serializer_class = PartnerSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        agent_id = self.request.query_params.get('agent_id')
        if agent_id:
            queryset = queryset.filter(agent_id=agent_id)
        return queryset

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # time.sleep(2)
        self.perform_destroy(instance)
        return Response(
            {"message": "partnerDeleted"},
            status=status.HTTP_204_NO_CONTENT
        )
    

    def update(self, request, *args, **kwargs):
        # time.sleep(2)
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data) # res.data
    

        



class RegisterView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Пользователь создан'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializers

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(
            {
                'message': 'groupAdded',
                'group': serializer.data
            },
            status=status.HTTP_201_CREATED
        )



class AgentViewSet(viewsets.ModelViewSet):
    queryset = Agent.objects.all().order_by('-pk')
    serializer_class = AgentSerializer
    permission_classes = [IsAuthenticated]

    def list(self, request, *args, **kwargs):
        # time.sleep(1)  # задержка 2 секунды
        return super().list(request, *args, **kwargs)

    
    def destroy(self, request, *args, **kwargs):
        # return Response(
        #         {'message': 'supplierHasSupplies'},
        #         status=status.HTTP_400_BAD_REQUEST
        #     )
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({'message': 'AgentDeleted'}, status=status.HTTP_200_OK)
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response({'data': serializer.data, "type": "success"}, status=status.HTTP_201_CREATED)

 


class EmployeeViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Employee.objects.all().order_by('-pk')
    serializer_class = EmployeeSerializer

    # Фильтрация и поиск
    filter_backends = [filters.SearchFilter]
    search_fields = ['name'] # сюда поля для поиска
    # ordering_fields = ['name', 'email', 'id']  # опцио

    def create(self, request, *args, **kwargs):
        # Можно вставить задержку или лог
        # time.sleep(2)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        return Response(
            {
                'message': 'employeeAdded',
                'employee': serializer.data
            },
            status=status.HTTP_201_CREATED
        )

    def destroy(self, request, *args, **kwargs):
        # time.sleep(2)
        # return Response(
        #         {'message': 'employeeNotDeleted'},
        #         status=status.HTTP_400_BAD_REQUEST
        #     )
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(status=status.HTTP_200_OK)
    
    def update(self, request, *args, **kwargs):
        # time.sleep(2)
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)
    

    @action(detail=False, methods=['get'], url_path='export_excel')
    def export_excel(self, request):
        # применяем поиск/фильтры, как и в list()
        queryset = self.filter_queryset(self.get_queryset())

        # создаем Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"

        # Заголовки
        headers = ["№", "Işgär"]
        ws.append(headers)

        # Строки
        for index, emp in enumerate(queryset, 1):
            ws.append([index, emp.name])

        # Автоширина колонок
        for i, column in enumerate(ws.columns, 1):
            max_len = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 2

        # Отдаём как файл
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=employees.xlsx'
        wb.save(response)
        return response


# @api_view(['GET', 'POST'])
# def suppliers_list(request):
#     if request.method == 'GET':
#         suppliers = Supplier.objects.all()
#         serializer = SupplierSerializer(suppliers, many=True)
#         return Response(serializer.data)
    
#     elif request.method == 'POST':
#         serializer = SupplierSerializer(data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




class MySecureView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if user.groups.filter(name="admin").exists():
            return Response({"authUser": f"{request.user}", "authGroup": "admin"})
        elif user.groups.filter(name="worker").exists():
            return Response({"authUser": f"{request.user}", "authGroup": "worker"})
        else:
            return Response({"message": "Нет доступа"}, status=403)
        



class AssignPartnersToAgentView(APIView):
    def post(self, request):
        # time.sleep(2)
        partners_id = request.data.get("partners_id")
        agent_id = request.data.get("igent_id")

        if agent_id is None:
            return Response({"error": "Invalid data"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            agent = Agent.objects.get(id=agent_id)
        except Agent.DoesNotExist:
            return Response({"error": "Agent not found"}, status=404)

        # Обнулить старые связи
        Partner.objects.filter(agent=agent).update(agent=None)

        # Назначить новые связи
        if isinstance(partners_id, list):
            Partner.objects.filter(id__in=partners_id).update(agent=agent)

        return Response({"message": "partnerSuccessUpdated"}, status=200)
    


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]


class BrandViewSet(viewsets.ModelViewSet):
    queryset = Brand.objects.all()
    serializer_class = BrandSerializer
    permission_classes = [IsAuthenticated]


class ModelViewSet(viewsets.ModelViewSet):
    queryset = Model.objects.all()
    serializer_class = ModelSerializer
    permission_classes = [IsAuthenticated]


class TagViewSet(viewsets.ModelViewSet):
    queryset = Tag.objects.all()
    serializer_class = TagSerializer
    permission_classes = [IsAuthenticated]


class ProductImageViewSet(viewsets.ModelViewSet):
    queryset = ProductImage.objects.all()
    serializer_class = ProductImageSerializer
    permission_classes = [IsAuthenticated]  




class ProductUnitViewSet(viewsets.ModelViewSet):
    queryset = ProductUnit.objects.all()
    serializer_class = ProductUnitSerializer
    permission_classes = [IsAuthenticated]  


    
    # def destroy(self, request, *args, **kwargs):
    #     # return Response(
    #     #         {'message': 'supplierHasSupplies'},
    #     #         status=status.HTTP_400_BAD_REQUEST
    #     #     )
    #     instance = self.get_object()
    #     self.perform_destroy(instance)
    #     return Response({'message': 'AgentDeleted'}, status=status.HTTP_200_OK)
    
    # def create(self, request, *args, **kwargs):
    #     serializer = self.get_serializer(data=request.data)
    #     serializer.is_valid(raise_exception=True)
    #     self.perform_create(serializer)
    #     return Response({'data': serializer.data, "type": "success"}, status=status.HTTP_201_CREATED)




















class UnitOfMeasurementViewSet(viewsets.ModelViewSet):
    queryset = UnitOfMeasurement.objects.all()
    serializer_class = UnitOfMeasurementSerializer

    def create(self, request, *args, **kwargs):
        # time.sleep(2)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        # time.sleep(2)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        # time.sleep(2)
        return super().partial_update(request, *args, **kwargs)
    


@require_GET
def check_name_unique(request):
    name = request.GET.get('name', '').strip()
    exists = Product.objects.filter(name__iexact=name).exists()
    return JsonResponse({'exists': exists})


# dlya poiska producta for free add
@api_view(["GET"])
def search_products(request):
    query = request.GET.get("q", "")
    results = Product.objects.annotate(
        similarity=TrigramSimilarity("name", query)
    ).filter(similarity__gt=0.1).order_by("-similarity")[:10]

    return Response(ProductSerializer(results, many=True).data)





class PriceChangeReportView(APIView):
    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if not start_date or not end_date:
            return Response({"detail": "start_date и end_date обязательны"}, status=status.HTTP_400_BAD_REQUEST)

        start = parse_date(start_date)
        end = parse_date(end_date)

        if not start or not end:
            return Response({"detail": "Неверный формат даты. Используйте YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)

        queryset = PriceChangeHistory.objects.filter(changed_at__date__range=(start, end)).select_related('product__base_unit')
        serializer = PriceChangeReportSerializer(queryset, many=True)
        return Response(serializer.data)
    




class PriceChangeExcelDownloadView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        if not start_date or not end_date:
            return HttpResponse("start_date и end_date обязательны", status=400)

        start = parse_date(start_date)
        end = parse_date(end_date)

        queryset = PriceChangeHistory.objects.filter(changed_at__date__range=(start, end)).select_related("product__base_unit")

        # Создаём Excel-файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Изменения цен"

        # Вставляем заголовок (например, на первую строку)
        header_text = f"Отчёт по изменениям цен с {start_date} по {end_date}"
        ws.merge_cells('A1:L1')  # Объединяем ячейки с A1 по L1 для заголовка
        cell = ws['A1']
        cell.value = header_text
        cell.font = Font(size=14, bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # Заголовки таблицы начинаются со второй строки
        headers = [
            "#", "Продукт", "Ед. изм.", "Старая цена", "Кол-во", "Сумма старая",
            "Новая цена", "Кол-во", "Сумма новая", "Прибыль", "Убыток", "Дата"
        ]
        ws.append([])  # Пустая строка (вторая)
        ws.append(headers)  # Третья строка — заголовки

        for idx, record in enumerate(queryset, start=1):
            old_total = float(record.old_price) * float(record.quantity_at_change)
            new_total = float(record.new_price) * float(record.quantity_at_change)
            profit = max(record.difference, 0)
            loss = min(record.difference, 0)

            ws.append([
                idx,
                record.product.name,
                record.product.base_unit.name,
                float(record.old_price),
                float(record.quantity_at_change),
                round(old_total, 2),
                float(record.new_price),
                float(record.quantity_at_change),
                round(new_total, 2),
                round(profit, 2),
                round(loss, 2),
                record.changed_at.strftime("%Y-%m-%d %H:%M"),
            ])

        # Автоширина столбцов
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        # Ответ как файл
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=price_changes_report.xlsx'
        wb.save(response)
        return response